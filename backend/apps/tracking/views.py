"""
Track & Trace API views.

Security architecture:
- All endpoints require JWT authentication
- Location submission: only authenticated drivers can submit their own location
- Live view: admin/manager can see all, drivers see only their own
- Route history: admin/manager can see all, drivers see only their own
- Rate limiting on all endpoints
- Input validation + anomaly detection
- Audit logging for session start/stop
"""
import logging
from django.utils import timezone
from django.db.models import Prefetch, Subquery, OuterRef
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.core.permissions import IsAdminOrManager
from .models import TrackingSession, LocationPoint
from .serializers import (
    TrackingSessionCreateSerializer,
    TrackingSessionSerializer,
    LocationPointCreateSerializer,
    LocationPointBatchSerializer,
    LocationPointSerializer,
    LiveVehicleSerializer,
)
from .security import (
    TrackingSubmitThrottle,
    TrackingReadThrottle,
    detect_teleportation,
    validate_location_bounds,
    get_client_ip,
    calculate_route_distance,
)

logger = logging.getLogger('tracking')


class TrackingSessionView(APIView):
    """
    Start/stop tracking sessions.
    
    POST: Start a new tracking session
    DELETE: Stop the current active session
    GET: Get the current user's active session
    """
    permission_classes = [IsAuthenticated]
    throttle_classes = [TrackingReadThrottle]

    def get(self, request):
        """Get the current user's active tracking session."""
        session = TrackingSession.objects.filter(
            user=request.user, is_active=True
        ).select_related('vehicle').first()
        
        if session:
            serializer = TrackingSessionSerializer(session)
            return Response(serializer.data)
        return Response({'active': False}, status=status.HTTP_200_OK)

    def post(self, request):
        """Start a new tracking session."""
        # End any existing active sessions for this user
        TrackingSession.objects.filter(
            user=request.user, is_active=True
        ).update(is_active=False, ended_at=timezone.now())
        
        serializer = TrackingSessionCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        vehicle_id = serializer.validated_data.get('vehicle_id')
        vehicle = None
        if vehicle_id:
            from apps.fleet.models import Vehicle
            vehicle = Vehicle.objects.filter(id=vehicle_id).first()
        
        # Auto-detect vehicle from driver profile
        if not vehicle and hasattr(request.user, 'driver_profile'):
            driver = request.user.driver_profile
            if driver and driver.voertuig:
                vehicle = driver.voertuig

        session = TrackingSession.objects.create(
            user=request.user,
            vehicle=vehicle,
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
            ip_address=get_client_ip(request),
        )
        
        logger.info(
            f"Tracking session started: user={request.user.username}, "
            f"session={session.id}, vehicle={vehicle}"
        )
        
        result = TrackingSessionSerializer(session)
        return Response(result.data, status=status.HTTP_201_CREATED)

    def delete(self, request):
        """Stop the current user's active tracking session."""
        updated = TrackingSession.objects.filter(
            user=request.user, is_active=True
        ).update(is_active=False, ended_at=timezone.now())
        
        if updated:
            logger.info(f"Tracking session stopped: user={request.user.username}")
            return Response({'stopped': True})
        return Response(
            {'detail': 'No active session found.'},
            status=status.HTTP_404_NOT_FOUND,
        )


class LocationSubmitView(APIView):
    """
    Submit GPS location points.
    
    POST: Submit a single location point or batch of points.
    
    Security:
    - Authenticated users only
    - User must have an active tracking session
    - Coordinates validated for realistic ranges
    - Anomaly detection (teleportation check)
    - Rate limited to prevent spam
    """
    permission_classes = [IsAuthenticated]
    throttle_classes = [TrackingSubmitThrottle]

    def post(self, request):
        """Submit location point(s)."""
        session = TrackingSession.objects.filter(
            user=request.user, is_active=True
        ).first()
        
        if not session:
            return Response(
                {'detail': 'No active tracking session. Start one first.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        # Check if it's a batch or single point
        is_batch = 'points' in request.data
        
        if is_batch:
            batch_serializer = LocationPointBatchSerializer(data=request.data)
            batch_serializer.is_valid(raise_exception=True)
            points_data = batch_serializer.validated_data['points']
        else:
            point_serializer = LocationPointCreateSerializer(data=request.data)
            point_serializer.is_valid(raise_exception=True)
            points_data = [point_serializer.validated_data]
        
        # Get the last known point for anomaly detection
        last_point = session.points.order_by('-recorded_at').first()
        
        created_points = []
        rejected_count = 0
        
        for point_data in points_data:
            lat = float(point_data['latitude'])
            lon = float(point_data['longitude'])
            
            # Security: validate location bounds
            if not validate_location_bounds(lat, lon):
                rejected_count += 1
                continue
            
            # Security: teleportation detection
            if last_point:
                is_teleport = detect_teleportation(
                    float(last_point.latitude), float(last_point.longitude),
                    last_point.recorded_at,
                    lat, lon, point_data['recorded_at'],
                )
                if is_teleport:
                    logger.warning(
                        f"Teleportation rejected: user={request.user.username}, "
                        f"session={session.id}"
                    )
                    rejected_count += 1
                    continue
            
            point = LocationPoint(
                session=session,
                latitude=point_data['latitude'],
                longitude=point_data['longitude'],
                accuracy=point_data.get('accuracy'),
                speed=point_data.get('speed'),
                heading=point_data.get('heading'),
                altitude=point_data.get('altitude'),
                recorded_at=point_data['recorded_at'],
            )
            created_points.append(point)
            last_point = point  # Use for next iteration's anomaly check
        
        # Bulk create for performance
        if created_points:
            LocationPoint.objects.bulk_create(created_points)
        
        return Response({
            'accepted': len(created_points),
            'rejected': rejected_count,
        }, status=status.HTTP_201_CREATED)


class LiveTrackingView(APIView):
    """
    Get real-time positions of all active tracking sessions.
    
    Admin/manager: see all active vehicles
    Chauffeur: see only own position
    
    Security:
    - Role-based data filtering
    - Minimal data exposure (no IP, user agent, etc.)
    """
    permission_classes = [IsAuthenticated]
    throttle_classes = [TrackingReadThrottle]

    def get(self, request):
        """Get all active vehicle positions."""
        user = request.user
        
        # Get active sessions with their latest point
        sessions_qs = TrackingSession.objects.filter(
            is_active=True
        ).select_related('vehicle', 'user', 'user__driver_profile')
        
        # Role-based filtering
        if user.rol == 'chauffeur':
            sessions_qs = sessions_qs.filter(user=user)
        
        results = []
        for session in sessions_qs:
            last_point = session.points.order_by('-recorded_at').first()
            if not last_point:
                continue
            
            # Get driver name
            user_name = ''
            if hasattr(session.user, 'driver_profile') and session.user.driver_profile:
                user_name = session.user.driver_profile.naam
            else:
                user_name = session.user.full_name or session.user.username
            
            results.append({
                'session_id': session.id,
                'user_name': user_name,
                'vehicle_id': session.vehicle_id,
                'vehicle_kenteken': session.vehicle.kenteken if session.vehicle else None,
                'vehicle_ritnummer': session.vehicle.ritnummer if session.vehicle else None,
                'vehicle_type': session.vehicle.type_wagen if session.vehicle else None,
                'latitude': float(last_point.latitude),
                'longitude': float(last_point.longitude),
                'speed': last_point.speed,
                'heading': last_point.heading,
                'accuracy': last_point.accuracy,
                'recorded_at': last_point.recorded_at,
                'is_active': True,
                'vehicle_status': 'driving' if last_point.speed and last_point.speed > 3 else ('driving' if session.is_active else 'parked'),
            })
        
        serializer = LiveVehicleSerializer(results, many=True)
        return Response(serializer.data)


class RouteHistoryView(APIView):
    """
    Get route history for a specific session or date range.
    
    Admin/manager: access all sessions
    Chauffeur: access only own sessions
    """
    permission_classes = [IsAuthenticated]
    throttle_classes = [TrackingReadThrottle]

    def get(self, request, session_id=None):
        """Get route history for a session."""
        user = request.user
        
        if session_id:
            try:
                session = TrackingSession.objects.select_related(
                    'vehicle', 'user', 'user__driver_profile'
                ).get(id=session_id)
            except TrackingSession.DoesNotExist:
                return Response(
                    {'detail': 'Session not found.'},
                    status=status.HTTP_404_NOT_FOUND,
                )
            
            # Security: enforce ownership for chauffeurs
            if user.rol == 'chauffeur' and session.user_id != user.id:
                return Response(
                    {'detail': 'Not authorized.'},
                    status=status.HTTP_403_FORBIDDEN,
                )
            
            points = session.points.order_by('recorded_at')
            distance = calculate_route_distance(points)
            
            return Response({
                'session': TrackingSessionSerializer(session).data,
                'points': LocationPointSerializer(points, many=True).data,
                'total_points': points.count(),
                'distance_km': distance,
            })
        
        # List sessions (most recent first)
        sessions_qs = TrackingSession.objects.select_related(
            'vehicle', 'user', 'user__driver_profile'
        ).order_by('-started_at')
        
        if user.rol == 'chauffeur':
            sessions_qs = sessions_qs.filter(user=user)
        
        # Optional filters
        vehicle_id = request.query_params.get('vehicle')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        
        if vehicle_id:
            sessions_qs = sessions_qs.filter(vehicle_id=vehicle_id)
        if date_from:
            sessions_qs = sessions_qs.filter(started_at__date__gte=date_from)
        if date_to:
            sessions_qs = sessions_qs.filter(started_at__date__lte=date_to)
        
        sessions_qs = sessions_qs[:50]  # Limit to 50 sessions
        
        serializer = TrackingSessionSerializer(sessions_qs, many=True)
        return Response(serializer.data)


class TrackingVehiclesView(APIView):
    """
    Get list of vehicles available for tracking assignment.
    Used in the vehicle monitor dropdown.
    """
    permission_classes = [IsAuthenticated]
    throttle_classes = [TrackingReadThrottle]

    def get(self, request):
        """Get vehicles for tracking assignment."""
        try:
            from apps.fleet.models import Vehicle
            vehicles = Vehicle.objects.all().order_by('kenteken').values(
                'id', 'kenteken', 'type_wagen', 'ritnummer',
            )
            return Response(list(vehicles))
        except Exception as e:
            logger.error(f"Failed to load tracking vehicles: {e}")
            return Response([], status=status.HTTP_200_OK)


class AssignedVehicleView(APIView):
    """
    Get the vehicle assigned to the current logged-in driver.
    Returns the vehicle linked via Driver.voertuig → Driver.gekoppelde_gebruiker.
    """
    permission_classes = [IsAuthenticated]
    throttle_classes = [TrackingReadThrottle]

    def get(self, request):
        """Get the assigned vehicle for the current user."""
        try:
            from apps.drivers.models import Driver
            driver = Driver.objects.select_related('voertuig').filter(
                gekoppelde_gebruiker=request.user
            ).first()

            if not driver or not driver.voertuig:
                return Response({'assigned': False}, status=status.HTTP_200_OK)

            vehicle = driver.voertuig
            return Response({
                'assigned': True,
                'vehicle': {
                    'id': str(vehicle.id),
                    'kenteken': vehicle.kenteken,
                    'type_wagen': vehicle.type_wagen,
                    'ritnummer': vehicle.ritnummer,
                },
                'driver_naam': driver.naam,
            })
        except Exception as e:
            logger.error(f"Failed to load assigned vehicle: {e}")
            return Response({'assigned': False}, status=status.HTTP_200_OK)


class TachographOverviewView(APIView):
    """
    GET /api/tracking/tachograph/?date=YYYY-MM-DD
    Returns tachograph trip data for all vehicles on the given date.
    """
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        from .tachograph_archive_service import fetch_live_tachograph_data
        from .tachograph_service import FMTrackError
        from datetime import date as date_type

        date_str = request.query_params.get('date')
        if not date_str:
            date_str = date_type.today().isoformat()

        # Validate date format
        try:
            from datetime import datetime
            datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            return Response(
                {'error': 'Ongeldig datumformaat. Gebruik YYYY-MM-DD.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            data = fetch_live_tachograph_data(date_str)
            return Response({'date': date_str, 'vehicles': data, 'count': len(data)})
        except FMTrackError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def _enrich_with_driver_overtime(self, vehicles, date_str):
        """
        Recalculate overtime using driver settings (standaard_begintijd, uren_per_dag, standaard_pauze)
        instead of the hardcoded 8-hour threshold from the tachograph service.
        Adds a calculation breakdown string to each vehicle.
        """
        from apps.drivers.models import Driver
        from datetime import datetime, timedelta
        from zoneinfo import ZoneInfo

        NL_TZ = ZoneInfo('Europe/Amsterdam')

        # Build plate -> driver lookup
        plate_lookup = {}
        for driver in Driver.objects.filter(auto_uren=True).exclude(tacho_kenteken=''):
            normalized = driver.tacho_kenteken.upper().replace('-', '').replace(' ', '')
            plate_lookup[normalized] = driver

        for vehicle in vehicles:
            plate = vehicle.get('plate_number', '')
            normalized_plate = plate.upper().replace('-', '').replace(' ', '')
            driver = plate_lookup.get(normalized_plate)

            if not driver:
                vehicle['overtime_calculation'] = None
                continue

            last_end_str = vehicle.get('last_end')
            if not last_end_str:
                vehicle['overtime_calculation'] = None
                continue

            try:
                tacho_end = datetime.fromisoformat(
                    last_end_str.replace('Z', '+00:00')
                ).astimezone(NL_TZ).time()
                tacho_start_str = vehicle.get('first_start', '')
                tacho_start = datetime.fromisoformat(
                    tacho_start_str.replace('Z', '+00:00')
                ).astimezone(NL_TZ).time() if tacho_start_str else None
            except (ValueError, AttributeError):
                vehicle['overtime_calculation'] = None
                continue

            # Use driver's start time if set, otherwise tachograph start
            start_time = driver.standaard_begintijd or tacho_start
            if not start_time:
                vehicle['overtime_calculation'] = None
                continue

            end_time = tacho_end
            pauze_minutes = driver.standaard_pauze or 30
            uren_per_dag = float(driver.uren_per_dag) if driver.uren_per_dag is not None else 8.0

            # Calculate total work hours
            from datetime import date as date_cls
            start_dt = datetime.combine(date_cls.today(), start_time)
            end_dt = datetime.combine(date_cls.today(), end_time)
            if end_dt < start_dt:
                end_dt += timedelta(days=1)

            total_work_seconds = (end_dt - start_dt).total_seconds()
            total_work_hours = total_work_seconds / 3600
            pauze_hours = pauze_minutes / 60
            netto_hours = total_work_hours - pauze_hours
            overtime = max(0, round(netto_hours - uren_per_dag, 2))

            # Update vehicle overtime fields
            vehicle['overtime_hours'] = overtime
            vehicle['has_overtime'] = overtime > 0

            def _fmt(h):
                hrs = int(h)
                mins = int(round((h - hrs) * 60))
                return f"{hrs:02d}:{mins:02d}"

            vehicle['overtime_display'] = _fmt(overtime) if overtime > 0 else None

            # Build calculation breakdown
            vehicle['overtime_calculation'] = {
                'driver_name': driver.naam,
                'start_time': start_time.strftime('%H:%M'),
                'end_time': end_time.strftime('%H:%M'),
                'total_work_hours': round(total_work_hours, 2),
                'total_work_display': _fmt(total_work_hours),
                'pauze_minutes': pauze_minutes,
                'pauze_display': _fmt(pauze_hours),
                'netto_hours': round(netto_hours, 2),
                'netto_display': _fmt(netto_hours),
                'uren_per_dag': uren_per_dag,
                'uren_per_dag_display': _fmt(uren_per_dag),
                'overtime_hours': overtime,
                'overtime_display': _fmt(overtime) if overtime > 0 else '00:00',
                'formula': f"{_fmt(netto_hours)} - {_fmt(uren_per_dag)} = {_fmt(overtime)} overuren",
            }


class TachographOvertimeWriteView(APIView):
    """
    POST /api/tracking/tachograph/overtime/
    Write overtime hours from tachograph data to a driver.

    Body: {
        "driver_id": "<TMS driver UUID>",
        "date": "YYYY-MM-DD",
        "overtime_hours": 1.5,
        "vehicle_name": "99-BRD-5",
        "fm_driver_name": "Jan Scherpenisse"
    }
    """
    permission_classes = [IsAdminOrManager]

    def post(self, request):
        from apps.drivers.models import Driver
        from apps.tracking.models import TachographOvertime

        driver_id = request.data.get('driver_id')
        date_str = request.data.get('date')
        overtime_hours = request.data.get('overtime_hours')
        vehicle_name = request.data.get('vehicle_name', '')
        fm_driver_name = request.data.get('fm_driver_name', '')

        if not all([driver_id, date_str, overtime_hours]):
            return Response(
                {'error': 'driver_id, date en overtime_hours zijn verplicht.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            driver = Driver.objects.get(pk=driver_id)
        except Driver.DoesNotExist:
            return Response(
                {'error': 'Chauffeur niet gevonden.'},
                status=status.HTTP_404_NOT_FOUND
            )

        overtime, created = TachographOvertime.objects.update_or_create(
            driver=driver,
            date=date_str,
            defaults={
                'overtime_hours': overtime_hours,
                'vehicle_name': vehicle_name,
                'fm_driver_name': fm_driver_name,
                'created_by': request.user,
            }
        )

        return Response({
            'success': True,
            'created': created,
            'message': f'Overuren ({overtime_hours}u) {"aangemaakt" if created else "bijgewerkt"} voor {driver.naam}.',
            'id': str(overtime.id),
        }, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)


class TachographOvertimeListView(APIView):
    """
    GET /api/tracking/tachograph/overtime/?driver_id=...&date_from=...&date_till=...
    List tachograph overtime records.
    """
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        from apps.tracking.models import TachographOvertime

        qs = TachographOvertime.objects.select_related('driver').order_by('-date')

        driver_id = request.query_params.get('driver_id')
        if driver_id:
            qs = qs.filter(driver_id=driver_id)

        date_from = request.query_params.get('date_from')
        if date_from:
            qs = qs.filter(date__gte=date_from)

        date_till = request.query_params.get('date_till')
        if date_till:
            qs = qs.filter(date__lte=date_till)

        records = []
        for ot in qs[:100]:
            records.append({
                'id': str(ot.id),
                'driver_id': str(ot.driver_id),
                'driver_naam': ot.driver.naam,
                'date': ot.date.isoformat(),
                'overtime_hours': float(ot.overtime_hours),
                'vehicle_name': ot.vehicle_name,
                'fm_driver_name': ot.fm_driver_name,
                'created_at': ot.created_at.isoformat(),
            })

        return Response({'results': records, 'count': len(records)})


class FMTrackPositionsView(APIView):
    """
    GET /api/tracking/fm-positions/
    Returns last known positions of all FM-Track vehicles.
    Matches FM-Track plate_number to TMS Vehicle kenteken for linking.
    """
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        from apps.tracking.tachograph_service import get_vehicle_locations, FMTrackError
        from apps.fleet.models import Vehicle

        try:
            positions = get_vehicle_locations()
        except FMTrackError as e:
            logger.warning('FM-Track positions unavailable: %s', e)
            return Response(
                {
                    'error': 'De FM-Track-service is tijdelijk niet beschikbaar. Probeer het later opnieuw.',
                    'code': 'fm_track_unavailable',
                },
                status=status.HTTP_502_BAD_GATEWAY,
            )
        except Exception:
            logger.exception('Unexpected error while loading FM-Track positions')
            return Response(
                {'error': 'FM-Track posities konden niet worden geladen.', 'code': 'fm_track_internal_error'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # Build a lookup of TMS vehicles by kenteken for enrichment
        vehicle_lookup = {}
        for v in Vehicle.objects.all():
            if v.kenteken:
                vehicle_lookup[v.kenteken.upper().replace('-', '')] = {
                    'id': str(v.id),
                    'kenteken': v.kenteken,
                    'ritnummer': v.ritnummer,
                }

        # Enrich positions with TMS vehicle data
        for pos in positions:
            plate = (pos.get('plate_number') or '').upper().replace('-', '')
            tms_vehicle = vehicle_lookup.get(plate)
            if tms_vehicle:
                pos['tms_vehicle'] = tms_vehicle
            else:
                pos['tms_vehicle'] = None

        return Response({'positions': positions, 'count': len(positions)})


class TachographVehiclesListView(APIView):
    """
    GET /api/tracking/tachograph/vehicles/
    Returns a list of all FM-Track vehicle plate numbers.
    Used for the tacho_kenteken dropdown in driver settings.
    """
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        from apps.tracking.tachograph_service import get_objects, FMTrackError

        try:
            objects = get_objects()
        except FMTrackError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        vehicles = []
        for obj in objects:
            vehicle_params = obj.get('vehicle_params', {})
            plate = vehicle_params.get('plate_number') or obj.get('name', '')
            vehicles.append({
                'object_id': obj.get('id', ''),
                'name': obj.get('name', ''),
                'plate_number': plate,
                'make': vehicle_params.get('make', ''),
                'model': vehicle_params.get('model', ''),
            })

        vehicles.sort(key=lambda x: x['plate_number'])
        return Response({'vehicles': vehicles, 'count': len(vehicles)})


class TachographManualSyncView(APIView):
    """
    GET  /api/tracking/tachograph/sync/  — Return sync info (start date, unprocessed count).
    POST /api/tracking/tachograph/sync/  — Trigger sync.

    Pass {"force": true} to clear existing sync data and re-sync everything.
    """
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        from datetime import date, timedelta
        from apps.core.models import AppSettings
        from apps.tracking.models import TachographSyncLog

        settings = AppSettings.get_settings()
        start_datum = settings.tachograaf_start_datum

        if not start_datum:
            return Response({'start_datum': None, 'effective_start': None, 'unprocessed_dates': 0})

        today = date.today()
        yesterday = today - timedelta(days=1)

        # FM-Track only allows the last 29 days
        MAX_HISTORY = 29
        earliest_allowed = today - timedelta(days=MAX_HISTORY)
        effective_start = max(start_datum, earliest_allowed)

        synced_dates = set(
            TachographSyncLog.objects.filter(
                date__gte=effective_start,
                date__lte=yesterday,
            ).values_list('date', flat=True)
        )

        unprocessed = 0
        current = effective_start
        while current <= yesterday:
            if current not in synced_dates:
                unprocessed += 1
            current += timedelta(days=1)

        return Response({
            'start_datum': start_datum.strftime('%Y-%m-%d'),
            'effective_start': effective_start.strftime('%Y-%m-%d'),
            'unprocessed_dates': unprocessed,
        })

    def post(self, request):
        force = request.data.get('force', False)

        try:
            if force:
                from apps.tracking.tasks import force_resync_tachograph_hours
                result = force_resync_tachograph_hours()
            else:
                from apps.tracking.tasks import sync_tachograph_hours
                result = sync_tachograph_hours()
            return Response(result)
        except Exception as e:
            logger.exception('Manual tachograph sync failed')
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class TachographArchiveListView(APIView):
    """
    GET /api/tracking/tachograph/archive/?date=YYYY-MM-DD
    Returns stored tachograph archive data from local database.
    """
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        from datetime import date as date_type
        from .tachograph_archive_service import get_tachograph_archive

        date_str = request.query_params.get('date')
        date_from = request.query_params.get('date_from')
        date_till = request.query_params.get('date_till')
        if not date_str and not date_from and not date_till:
            date_str = date_type.today().isoformat()

        from datetime import datetime
        try:
            if date_str:
                datetime.strptime(date_str, '%Y-%m-%d')
            if date_from:
                datetime.strptime(date_from, '%Y-%m-%d')
            if date_till:
                datetime.strptime(date_till, '%Y-%m-%d')
        except ValueError:
            return Response(
                {'error': 'Ongeldig datumformaat. Gebruik YYYY-MM-DD.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rows = get_tachograph_archive(
            date_str=date_str,
            date_from=date_from,
            date_till=date_till,
        )
        return Response({
            'date': date_str,
            'vehicles': rows,
            'count': len(rows),
        })


class TachographArchiveSyncView(APIView):
    """
    POST /api/tracking/tachograph/archive/sync/
    Body: {"date": "YYYY-MM-DD"}
    """
    permission_classes = [IsAdminOrManager]

    def post(self, request):
        from .tachograph_archive_service import upsert_tachograph_archive_for_date
        from .tachograph_service import FMTrackError

        date_str = request.data.get('date')
        if not date_str:
            return Response(
                {'error': 'Veld "date" is verplicht (YYYY-MM-DD).'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            result = upsert_tachograph_archive_for_date(date_str)
            return Response(result, status=status.HTTP_200_OK)
        except ValueError:
            return Response(
                {'error': 'Ongeldig datumformaat. Gebruik YYYY-MM-DD.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except FMTrackError:
            return Response(
                {'error': 'FM-Track gegevens konden niet worden opgehaald voor deze dag.'},
                status=status.HTTP_502_BAD_GATEWAY,
            )
        except Exception:
            logger.exception('Tachograaf archief sync mislukt voor %s', date_str)
            return Response(
                {'error': 'Synchronisatie van tachograaf archief is mislukt.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class TachographComparisonView(APIView):
    """
    GET /api/tracking/tachograph/comparison/?date_from=YYYY-MM-DD&date_till=YYYY-MM-DD
    Compare tachograph hours with submitted TimeEntry hours for a date range.
    Fetches tachograph data live from FM-Track API per relevant plate.
    Optional: &format=xlsx to download as Excel.
    """
    permission_classes = [IsAdminOrManager]

    def _build_rows(self, request):
        """Build comparison rows and return (rows, date_from_str, date_till_str) or Response on error."""
        from datetime import datetime as dt, timedelta
        from zoneinfo import ZoneInfo
        from collections import defaultdict
        from concurrent.futures import ThreadPoolExecutor, as_completed
        from apps.timetracking.models import TimeEntry
        from apps.drivers.models import Driver
        from apps.tracking.tachograph_service import (
            get_objects, get_drivers, get_trips,
            _build_vehicle_summary, _format_driver_name, FMTrackError,
        )

        NL_TZ = ZoneInfo('Europe/Amsterdam')
        UTC = ZoneInfo('UTC')

        date_from_str = request.query_params.get('date_from')
        date_till_str = request.query_params.get('date_till')
        if not date_from_str or not date_till_str:
            return Response(
                {'error': 'Parameters date_from en date_till zijn verplicht (YYYY-MM-DD).'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            from_date = dt.strptime(date_from_str, '%Y-%m-%d').date()
            till_date = dt.strptime(date_till_str, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {'error': 'Ongeldig datumformaat. Gebruik YYYY-MM-DD.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if from_date > till_date:
            return Response(
                {'error': 'date_from mag niet na date_till liggen.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if (till_date - from_date).days > 31:
            return Response(
                {'error': 'Maximaal 31 dagen per keer.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        def _norm(plate):
            return (plate or '').upper().replace('-', '').replace(' ', '')

        # ── 1. FM-Track objects & drivers (2 API calls) ──
        try:
            fm_objects = get_objects()
            fm_drivers_list = get_drivers()
        except FMTrackError as e:
            return Response(
                {'error': f'FM-Track fout: {e}'},
                status=status.HTTP_502_BAD_GATEWAY,
            )

        driver_map = {d['id']: _format_driver_name(d) for d in fm_drivers_list}

        # plate_norm → FM-Track object
        plate_to_obj = {}
        for obj in fm_objects:
            plate = (
                obj.get('vehicle_params', {}).get('plate_number')
                or obj.get('name', '')
            )
            n = _norm(plate)
            if n:
                plate_to_obj[n] = obj

        # ── 2. TMS driver→plate mapping + user→driver mapping ──
        all_drivers = list(
            Driver.objects.select_related('gekoppelde_gebruiker').all()
        )
        plate_driver_map = {}
        user_driver_map = {}   # user_id → Driver
        for drv in all_drivers:
            if drv.tacho_kenteken:
                plate_driver_map[_norm(drv.tacho_kenteken)] = drv
            if drv.gekoppelde_gebruiker_id:
                user_driver_map[drv.gekoppelde_gebruiker_id] = drv

        # ── 3. All time entries in the range (concept + ingediend) ──
        all_entries = list(
            TimeEntry.objects.select_related('user').filter(
                datum__range=[from_date, till_date],
                status__in=['concept', 'ingediend'],
            )
        )

        # Optional kenteken filter from frontend
        kenteken_filter = request.query_params.get('kenteken', '').strip()
        kenteken_filter_norm = _norm(kenteken_filter) if kenteken_filter else ''

        entries_by_date_plate = {}      # manual/chauffeur entries
        auto_entries_by_date_plate = {}  # auto-import entries
        entries_by_date_user = {}   # (date, user_id) → entry (for drivers without plate)
        entry_plates = set()
        for entry in all_entries:
            is_auto = getattr(entry, 'bron', 'handmatig') == 'auto_import'

            # Index by user for drivers without tacho_kenteken
            ukey = (entry.datum, entry.user_id)
            if not is_auto and ukey not in entries_by_date_user:
                entries_by_date_user[ukey] = entry

            if not entry.kenteken:
                continue
            n = _norm(entry.kenteken)
            key = (entry.datum, n)

            if is_auto:
                if key not in auto_entries_by_date_plate:
                    auto_entries_by_date_plate[key] = entry
            else:
                if key not in entries_by_date_plate:
                    entries_by_date_plate[key] = entry
                else:
                    tms_drv = plate_driver_map.get(n)
                    if tms_drv and tms_drv.gekoppelde_gebruiker_id == entry.user_id:
                        entries_by_date_plate[key] = entry
            entry_plates.add(n)

        # ── 4. Determine which FM-Track vehicles to fetch ──
        if kenteken_filter_norm and kenteken_filter_norm in plate_to_obj:
            # Only fetch the selected vehicle for efficiency
            objs_to_fetch = {kenteken_filter_norm: plate_to_obj[kenteken_filter_norm]}
        else:
            objs_to_fetch = dict(plate_to_obj)  # fetch every vehicle from the API

        # ── 5. Fetch trips per vehicle for the full range (parallel) ──
        utc_start = dt(
            from_date.year, from_date.month, from_date.day, 0, 0, 0,
            tzinfo=NL_TZ,
        ).astimezone(UTC)
        utc_end = dt(
            till_date.year, till_date.month, till_date.day, 23, 59, 59,
            tzinfo=NL_TZ,
        ).astimezone(UTC)

        tacho_by_date_plate = {}  # (date, plate_norm) → vehicle_summary

        def _fetch_and_group(plate, obj):
            try:
                trips = get_trips(obj['id'], utc_start, utc_end)
                return plate, obj, (trips or [])
            except FMTrackError:
                return plate, obj, []

        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = [
                executor.submit(_fetch_and_group, p, o)
                for p, o in objs_to_fetch.items()
            ]
            for future in as_completed(futures):
                plate, obj, trips = future.result()
                if not trips:
                    continue
                trips_by_date = defaultdict(list)
                for trip in trips:
                    ts = trip.get('trip_start', {}).get('datetime', '')
                    if ts:
                        try:
                            td = dt.strptime(ts[:10], '%Y-%m-%d').date()
                            if from_date <= td <= till_date:
                                trips_by_date[td].append(trip)
                        except ValueError:
                            pass
                for d, d_trips in trips_by_date.items():
                    summary = _build_vehicle_summary(obj, d_trips, driver_map)
                    tacho_by_date_plate[(d, plate)] = summary

        # ── 6. Build comparison rows ──
        rows = []
        current = from_date
        while current <= till_date:
            tacho_plates_today = {p for (d, p) in tacho_by_date_plate if d == current}
            entry_plates_today = {p for (d, p) in entries_by_date_plate if d == current}
            all_plates_today = tacho_plates_today | entry_plates_today

            # When a kenteken filter is active, always include that plate
            # so we generate a row even if no trips/entries exist
            if kenteken_filter_norm:
                all_plates_today.add(kenteken_filter_norm)

            # Track which users already appeared via plate-based rows
            seen_users_today = set()

            for norm_plate in sorted(all_plates_today):
                vehicle = tacho_by_date_plate.get((current, norm_plate))
                entry = entries_by_date_plate.get((current, norm_plate))
                auto_entry = auto_entries_by_date_plate.get((current, norm_plate))
                tms_driver = plate_driver_map.get(norm_plate)

                raw_kenteken = ''
                if vehicle:
                    raw_kenteken = vehicle.get(
                        'plate_number', vehicle.get('vehicle_name', ''),
                    )
                elif entry:
                    raw_kenteken = entry.kenteken
                elif auto_entry:
                    raw_kenteken = auto_entry.kenteken
                elif kenteken_filter_norm:
                    # Filtered plate with no trips/entries: use original plate from FM-Track
                    fm_obj = plate_to_obj.get(norm_plate)
                    if fm_obj:
                        raw_kenteken = (
                            fm_obj.get('vehicle_params', {}).get('plate_number')
                            or fm_obj.get('name', '')
                        )
                    else:
                        raw_kenteken = kenteken_filter

                driver_naam = tms_driver.naam if tms_driver else ''
                if not driver_naam and entry and entry.user:
                    driver_naam = entry.user.full_name
                if not driver_naam and auto_entry and auto_entry.user:
                    driver_naam = auto_entry.user.full_name

                if entry and entry.user_id:
                    seen_users_today.add(entry.user_id)
                if auto_entry and auto_entry.user_id:
                    seen_users_today.add(auto_entry.user_id)
                if tms_driver and tms_driver.gekoppelde_gebruiker_id:
                    seen_users_today.add(tms_driver.gekoppelde_gebruiker_id)

                rows.append(self._build_single_row(
                    current, raw_kenteken, driver_naam, vehicle, entry,
                    auto_entry, tms_driver, NL_TZ,
                ))

            # Also include drivers who have entries today but no plate match
            entry_users_today = {
                uid for (d, uid) in entries_by_date_user if d == current
            }
            for user_id in sorted(entry_users_today - seen_users_today):
                entry = entries_by_date_user.get((current, user_id))
                if not entry:
                    continue
                tms_driver = user_driver_map.get(user_id)
                driver_naam = tms_driver.naam if tms_driver else ''
                if not driver_naam and entry.user:
                    driver_naam = entry.user.full_name
                raw_kenteken = entry.kenteken or ''

                rows.append(self._build_single_row(
                    current, raw_kenteken, driver_naam, None, entry,
                    None, tms_driver, NL_TZ,
                ))

            current += timedelta(days=1)

        # ── 7. Build drivers list for frontend filter ──
        drivers_list = [
            {'id': str(drv.id), 'naam': drv.naam}
            for drv in sorted(all_drivers, key=lambda d: d.naam)
        ]

        # ── 8. Build plates list for frontend filter (all from Linqo API) ──
        plates_list = []
        for obj in fm_objects:
            plate = (
                obj.get('vehicle_params', {}).get('plate_number')
                or obj.get('name', '')
            )
            if plate:
                plates_list.append(plate)
        plates_list = sorted(set(plates_list))

        return rows, date_from_str, date_till_str, drivers_list, plates_list

    @staticmethod
    def _build_single_row(current, raw_kenteken, driver_naam, vehicle, entry, auto_entry, tms_driver, NL_TZ):
        from datetime import datetime as dt

        # Tachograph times
        tacho_begin = tacho_eind = None
        tacho_total = None
        if vehicle:
            fs = vehicle.get('first_start')
            le = vehicle.get('last_end')
            fs_dt = le_dt = None
            if fs:
                try:
                    fs_dt = dt.fromisoformat(
                        fs.replace('Z', '+00:00'),
                    ).astimezone(NL_TZ)
                    tacho_begin = fs_dt.strftime('%H:%M')
                except (ValueError, AttributeError):
                    pass
            if le:
                try:
                    le_dt = dt.fromisoformat(
                        le.replace('Z', '+00:00'),
                    ).astimezone(NL_TZ)
                    tacho_eind = le_dt.strftime('%H:%M')
                except (ValueError, AttributeError):
                    pass
            if fs_dt and le_dt:
                tacho_total = round(
                    (le_dt - fs_dt).total_seconds() / 3600, 2,
                )

        # Chauffeur times (manual entries)
        chauffeur_begin = chauffeur_eind = None
        chauffeur_total = None
        if entry:
            if entry.aanvang:
                chauffeur_begin = entry.aanvang.strftime('%H:%M')
            if entry.eind:
                chauffeur_eind = entry.eind.strftime('%H:%M')
            if entry.totaal_uren:
                chauffeur_total = round(
                    entry.totaal_uren.total_seconds() / 3600, 2,
                )

        # Auto-import times
        auto_begin = auto_eind = None
        auto_totaal = None
        if auto_entry:
            if auto_entry.aanvang:
                auto_begin = auto_entry.aanvang.strftime('%H:%M')
            if auto_entry.eind:
                auto_eind = auto_entry.eind.strftime('%H:%M')
            if auto_entry.totaal_uren:
                auto_totaal = round(
                    auto_entry.totaal_uren.total_seconds() / 3600, 2,
                )

        # Difference
        verschil = None
        verschil_bron = None
        if chauffeur_total is not None and tacho_total:
            diff = round(chauffeur_total - tacho_total, 2)
            if diff != 0:
                verschil = abs(diff)
                verschil_bron = 'chauffeur' if diff > 0 else 'tacho'

        # Uren per dag + overwerk
        uren_per_dag = None
        overwerk_uren = None
        overwerk_tacho = None
        drv_upd = getattr(tms_driver, 'uren_per_dag', None) if tms_driver else None
        if drv_upd is not None:
            upd = float(drv_upd)
            uren_per_dag = upd
            if chauffeur_total is not None and chauffeur_total > upd:
                overwerk_uren = round(chauffeur_total - upd, 2)
            if tacho_total is not None and tacho_total > upd:
                overwerk_tacho = round(tacho_total - upd, 2)

        return {
            'datum': current.strftime('%Y-%m-%d'),
            'kenteken': raw_kenteken,
            'chauffeur_naam': driver_naam,
            'chauffeur_begin': chauffeur_begin,
            'chauffeur_eind': chauffeur_eind,
            'tacho_begin': tacho_begin,
            'tacho_eind': tacho_eind,
            'chauffeur_totaal': chauffeur_total,
            'tacho_totaal': tacho_total,
            'auto_begin': auto_begin,
            'auto_eind': auto_eind,
            'auto_totaal': auto_totaal,
            'verschil': verschil,
            'verschil_bron': verschil_bron,
            'uren_per_dag': uren_per_dag,
            'overwerk_uren': overwerk_uren,
            'overwerk_tacho': overwerk_tacho,
        }

    def get(self, request):
        try:
            result = self._build_rows(request)
        except Exception as exc:
            import logging
            logging.getLogger(__name__).exception('TachographComparison error')
            return Response(
                {'error': f'Interne fout: {exc}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        if isinstance(result, Response):
            return result

        rows, date_from_str, date_till_str, drivers_list, plates_list = result

        # ── Export ──
        fmt = request.query_params.get('export')
        if fmt == 'xlsx':
            return self._export_xlsx(rows, date_from_str, date_till_str)
        if fmt == 'pdf':
            return self._export_pdf(rows, date_from_str, date_till_str)

        return Response({
            'date_from': date_from_str,
            'date_till': date_till_str,
            'rows': rows,
            'count': len(rows),
            'drivers': drivers_list,
            'plates': plates_list,
        })

    @staticmethod
    def _format_hours(hours):
        if hours is None:
            return ''
        h = int(hours)
        m = round((hours - h) * 60)
        return f'{h}:{m:02d}'

    def _export_xlsx(self, rows, date_from_str, date_till_str):
        import io
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = 'Uren Vergelijking'

        # Styles
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        odd_fill = PatternFill(start_color='EBF1F8', end_color='EBF1F8', fill_type='solid')
        thin_border = Border(
            bottom=Side(style='thin', color='D0D5DD'),
        )
        center_align = Alignment(horizontal='center', vertical='center')

        headers = [
            'Datum', 'Chauffeur', 'Kenteken',
            'Begin (uren)', 'Eind (uren)',
            'Aut. Begin', 'Aut. Eind',
            'Begin (tacho)', 'Eind (tacho)',
            'Totaal (uren)', 'Totaal (tacho)', 'Verschil',
            'Uren/dag', 'Overwerk (uren)', 'Overwerk (tacho)',
        ]
        col_widths = [14, 20, 14, 13, 13, 11, 11, 13, 13, 14, 14, 16, 12, 16, 16]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            ws.column_dimensions[cell.column_letter].width = width

        for row_idx, row in enumerate(rows, 2):
            is_odd = (row_idx % 2) == 0  # data row 1 = even excel row

            verschil_text = ''
            if row['verschil'] is not None:
                bron = 'Chauffeur' if row['verschil_bron'] == 'chauffeur' else 'Tacho'
                verschil_text = f'+{self._format_hours(row["verschil"])} ({bron})'

            values = [
                row['datum'],
                row['chauffeur_naam'],
                row['kenteken'],
                row['chauffeur_begin'] or '',
                row['chauffeur_eind'] or '',
                row.get('auto_begin') or '',
                row.get('auto_eind') or '',
                row['tacho_begin'] or '',
                row['tacho_eind'] or '',
                self._format_hours(row['chauffeur_totaal']),
                self._format_hours(row['tacho_totaal']),
                verschil_text,
                self._format_hours(row.get('uren_per_dag')),
                self._format_hours(row.get('overwerk_uren')),
                self._format_hours(row.get('overwerk_tacho')),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center_align if col_idx >= 4 else Alignment(vertical='center')
                cell.border = thin_border
                if is_odd:
                    cell.fill = odd_fill

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'uren_vergelijking_{date_from_str}_{date_till_str}.xlsx'
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _export_pdf(self, rows, date_from_str, date_till_str):
        import io
        from django.http import HttpResponse
        from weasyprint import HTML

        # Build HTML table
        html_rows = ''
        for i, row in enumerate(rows):
            bg = ' style="background:#EBF1F8"' if i % 2 == 0 else ''
            verschil_text = ''
            if row['verschil'] is not None:
                bron = 'Chauffeur' if row['verschil_bron'] == 'chauffeur' else 'Tacho'
                verschil_text = f'+{self._format_hours(row["verschil"])} ({bron})'
            html_rows += f'''<tr{bg}>
                <td>{row["datum"]}</td>
                <td>{row["chauffeur_naam"]}</td>
                <td>{row["kenteken"]}</td>
                <td class="c">{row["chauffeur_begin"] or "-"}</td>
                <td class="c">{row["chauffeur_eind"] or "-"}</td>
                <td class="c">{row.get("auto_begin") or "-"}</td>
                <td class="c">{row.get("auto_eind") or "-"}</td>
                <td class="c">{row["tacho_begin"] or "-"}</td>
                <td class="c">{row["tacho_eind"] or "-"}</td>
                <td class="c">{self._format_hours(row["chauffeur_totaal"]) or "-"}</td>
                <td class="c">{self._format_hours(row["tacho_totaal"]) or "-"}</td>
                <td class="c">{verschil_text or "\u2713"}</td>
                <td class="c">{self._format_hours(row.get("uren_per_dag")) or "-"}</td>
                <td class="c">{self._format_hours(row.get("overwerk_uren")) or "-"}</td>
                <td class="c">{self._format_hours(row.get("overwerk_tacho")) or "-"}</td>
            </tr>'''

        html_content = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: A4 landscape; margin: 15mm; }}
  body {{ font-family: Arial, sans-serif; font-size: 10px; }}
  h1 {{ font-size: 16px; color: #1F4E79; margin-bottom: 4px; }}
  .sub {{ font-size: 11px; color: #666; margin-bottom: 12px; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th {{ background: #1F4E79; color: #fff; padding: 6px 4px; font-size: 9px;
       text-transform: uppercase; text-align: left; }}
  td {{ padding: 5px 4px; border-bottom: 1px solid #ddd; }}
  .c {{ text-align: center; }}
  th.c {{ text-align: center; }}
</style></head><body>
<h1>Uren Vergelijking</h1>
<div class="sub">{date_from_str} &mdash; {date_till_str}</div>
<table>
  <thead><tr>
    <th>Datum</th><th>Chauffeur</th><th>Kenteken</th>
    <th class="c">Begin (uren)</th><th class="c">Eind (uren)</th>
    <th class="c">Aut. Begin</th><th class="c">Aut. Eind</th>
    <th class="c">Begin (tacho)</th><th class="c">Eind (tacho)</th>
    <th class="c">Totaal (uren)</th><th class="c">Totaal (tacho)</th>
    <th class="c">Verschil</th>
    <th class="c">Uren/dag</th><th class="c">Overwerk (uren)</th><th class="c">Overwerk (tacho)</th>
  </tr></thead>
  <tbody>{html_rows}</tbody>
</table>
</body></html>'''

        pdf = HTML(string=html_content).write_pdf()
        filename = f'uren_vergelijking_{date_from_str}_{date_till_str}.pdf'
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class VehicleDetailView(APIView):
    """
    GET /api/tracking/fm-positions/<object_id>/detail/?date=YYYY-MM-DD
    Returns detailed trip history for a specific FM-Track vehicle on a given date.
    Includes: current position, odometer, speed, trips with addresses, max speed per trip.
    """
    permission_classes = [IsAdminOrManager]

    def get(self, request, object_id):
        from datetime import datetime as dt, date as date_type
        from apps.tracking.tachograph_service import (
            get_trips, get_raw_data, get_objects, get_vehicle_locations, FMTrackError,
            _format_address, _format_duration,
        )

        date_str = request.query_params.get('date')
        if not date_str:
            date_str = date_type.today().isoformat()

        try:
            dt.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            return Response(
                {'error': 'Ongeldig datumformaat. Gebruik YYYY-MM-DD.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        target_date = dt.strptime(date_str, '%Y-%m-%d').date()
        date_from = dt(target_date.year, target_date.month, target_date.day, 0, 0, 0)
        date_till = dt(target_date.year, target_date.month, target_date.day, 23, 59, 59)

        try:
            # Fetch trips for this vehicle on this date
            trips_raw = get_trips(object_id, date_from, date_till)

            # Filter to target date
            trips_filtered = []
            for trip in trips_raw:
                trip_start_dt = trip.get('trip_start', {}).get('datetime', '')
                if trip_start_dt and trip_start_dt[:10] == date_str:
                    trips_filtered.append(trip)

            # Get current live position for this vehicle
            current_position = None
            try:
                positions = get_vehicle_locations()
                for pos in positions:
                    if pos.get('object_id') == object_id:
                        current_position = pos
                        break
            except FMTrackError:
                pass

            # Get vehicle info from objects
            vehicle_info = None
            fuel_tank_capacity = 0
            avg_fuel_consumption = 0
            try:
                objects = get_objects()
                for obj in objects:
                    if obj.get('id') == object_id:
                        vp = obj.get('vehicle_params', {})
                        vehicle_info = {
                            'name': obj.get('name', ''),
                            'plate_number': vp.get('plate_number') or obj.get('name', ''),
                            'make': vp.get('make', ''),
                            'model': vp.get('model', ''),
                        }
                        fuel_tank_capacity = vp.get('fuel_tank_capacity') or 0
                        avg_fuel_consumption = vp.get('average_fuel_consumption') or 0
                        break
            except FMTrackError:
                pass

            # Build trip details
            trip_details = []
            fuel_chart_data = []
            total_distance = 0
            total_duration_seconds = 0
            max_speed_overall = 0

            for trip in trips_filtered:
                start_info = trip.get('trip_start', {})
                end_info = trip.get('trip_end', {})
                duration = trip.get('trip_duration', 0)
                total_duration_seconds += duration

                start_mileage = start_info.get('mileage', 0)
                end_mileage = end_info.get('mileage', 0)
                distance_km = round(trip.get('mileage', 0) / 1000, 1)
                total_distance += distance_km

                # FM-Track provides max_speed in trip data (m/s or km/h depending on API version)
                trip_max_speed = trip.get('max_speed', 0)
                # Convert from m/s to km/h if needed (FM-Track v1 uses km/h)
                if trip_max_speed > max_speed_overall:
                    max_speed_overall = trip_max_speed

                # Speed limit check: >130 km/h is considered speeding in NL
                is_speeding = trip_max_speed > 130

                # Fuel levels from trip start/end
                start_fuel = start_info.get('fuel_level', 0) or 0
                end_fuel = end_info.get('fuel_level', 0) or 0

                # Calculate fuel in liters
                start_fuel_liters = round(fuel_tank_capacity * start_fuel / 100, 1) if fuel_tank_capacity and start_fuel else None
                end_fuel_liters = round(fuel_tank_capacity * end_fuel / 100, 1) if fuel_tank_capacity and end_fuel else None

                trip_details.append({
                    'start_time': start_info.get('datetime', ''),
                    'end_time': end_info.get('datetime', ''),
                    'start_address': _format_address(start_info.get('address')),
                    'end_address': _format_address(end_info.get('address')),
                    'start_lat': start_info.get('latitude'),
                    'start_lng': start_info.get('longitude'),
                    'end_lat': end_info.get('latitude'),
                    'end_lng': end_info.get('longitude'),
                    'start_km': round(start_mileage / 1000, 1) if start_mileage else 0,
                    'end_km': round(end_mileage / 1000, 1) if end_mileage else 0,
                    'distance_km': distance_km,
                    'duration_seconds': duration,
                    'duration_display': _format_duration(duration),
                    'max_speed': round(trip_max_speed),
                    'is_speeding': is_speeding,
                    'start_fuel_level': round(start_fuel),
                    'end_fuel_level': round(end_fuel),
                    'fuel_used_liters': round(start_fuel_liters - end_fuel_liters, 1) if start_fuel_liters is not None and end_fuel_liters is not None else None,
                })

                # Build fuel chart data points (timestamp + fuel %)
                if start_fuel > 0 and start_info.get('datetime'):
                    fuel_chart_data.append({
                        'timestamp': start_info['datetime'],
                        'fuel_level': round(start_fuel),
                        'fuel_liters': start_fuel_liters,
                        'event': 'trip_start',
                    })
                if end_fuel > 0 and end_info.get('datetime'):
                    fuel_chart_data.append({
                        'timestamp': end_info['datetime'],
                        'fuel_level': round(end_fuel),
                        'fuel_liters': end_fuel_liters,
                        'event': 'trip_end',
                    })

            # Current odometer from last trip or position
            current_km = 0
            if trip_details:
                current_km = trip_details[-1]['end_km']
            elif current_position and current_position.get('mileage'):
                current_km = round(current_position['mileage'] / 1000, 1)

            # Fuel consumption stats
            total_fuel_used_liters = None
            total_fuel_used_pct = None
            avg_fuel_consumption = None  # L/100km
            avg_fuel_consumption_pct = None  # %/100km
            fuel_per_hour = None
            fuel_per_hour_pct = None
            if fuel_chart_data and len(fuel_chart_data) >= 2:
                first_fuel = fuel_chart_data[0].get('fuel_level', 0)
                last_fuel = fuel_chart_data[-1].get('fuel_level', 0)
                if first_fuel > 0:
                    fuel_diff_pct = max(0, first_fuel - last_fuel)
                    total_fuel_used_pct = round(fuel_diff_pct, 1)
                    if fuel_tank_capacity > 0:
                        total_fuel_used_liters = round(fuel_tank_capacity * fuel_diff_pct / 100, 1)
                    if total_distance > 0 and fuel_diff_pct > 0:
                        avg_fuel_consumption_pct = round(fuel_diff_pct / total_distance * 100, 1)
                        if fuel_tank_capacity > 0 and total_fuel_used_liters:
                            avg_fuel_consumption = round(total_fuel_used_liters / total_distance * 100, 1)
                    if total_duration_seconds > 0 and fuel_diff_pct > 0:
                        fuel_per_hour_pct = round(fuel_diff_pct / (total_duration_seconds / 3600), 1)
                        if fuel_tank_capacity > 0 and total_fuel_used_liters:
                            fuel_per_hour = round(total_fuel_used_liters / (total_duration_seconds / 3600), 1)

            # Remaining driving time (EU 561/2006: max 9h daily)
            max_daily_drive_seconds = 9 * 3600
            remaining_drive_seconds = max(0, max_daily_drive_seconds - total_duration_seconds)

            # Fetch GPS route track for the day (raw_data gives actual GPS points)
            route_coordinates = []
            if trips_filtered:
                try:
                    raw_points = get_raw_data(object_id, date_from, date_till)
                    for pt in raw_points:
                        lat = pt.get('latitude') or pt.get('lat')
                        lng = pt.get('longitude') or pt.get('lng') or pt.get('lon')
                        if lat and lng:
                            route_coordinates.append({
                                'lat': float(lat),
                                'lng': float(lng),
                                'speed': pt.get('speed', 0),
                                'timestamp': pt.get('datetime') or pt.get('timestamp', ''),
                            })
                except (FMTrackError, Exception) as e:
                    logger.warning('Failed to fetch raw GPS data for route: %s', e)
                    # Fallback: build route from trip start/end coordinates
                    for trip_d in trip_details:
                        if trip_d.get('start_lat') and trip_d.get('start_lng'):
                            route_coordinates.append({
                                'lat': float(trip_d['start_lat']),
                                'lng': float(trip_d['start_lng']),
                                'speed': 0,
                                'timestamp': trip_d.get('start_time', ''),
                            })
                        if trip_d.get('end_lat') and trip_d.get('end_lng'):
                            route_coordinates.append({
                                'lat': float(trip_d['end_lat']),
                                'lng': float(trip_d['end_lng']),
                                'speed': 0,
                                'timestamp': trip_d.get('end_time', ''),
                            })

            # Build current_position with smart speed
            cur_pos = None
            if current_position:
                live_speed = current_position.get('speed', 0) or 0
                vehicle_status = current_position.get('vehicle_status', 'parked')
                speed_source = 'live'

                # When live GPS speed is 0 but vehicle is driving (e.g. between
                # GPS polls, at traffic light), derive a useful speed value from
                # the most recent trip so the UI isn't contradictory.
                if live_speed == 0 and vehicle_status == 'driving' and trip_details:
                    last_trip = trips_filtered[-1] if trips_filtered else None
                    if last_trip:
                        # Use average speed from last trip
                        lt_dur = last_trip.get('trip_duration', 0)
                        lt_dist = last_trip.get('mileage', 0) / 1000  # m -> km
                        if lt_dur > 0 and lt_dist > 0:
                            live_speed = round(lt_dist / (lt_dur / 3600))
                            speed_source = 'trip_avg'

                # If still 0 but driving, fall back to day average
                if live_speed == 0 and vehicle_status == 'driving' and total_distance > 0 and total_duration_seconds > 0:
                    live_speed = round(total_distance / (total_duration_seconds / 3600))
                    speed_source = 'day_avg'

                cur_pos = {
                    'latitude': current_position.get('latitude'),
                    'longitude': current_position.get('longitude'),
                    'speed': live_speed,
                    'speed_source': speed_source,
                    'address': current_position.get('address', ''),
                    'vehicle_status': vehicle_status,
                    'fuel_level': current_position.get('fuel_level', 0),
                    'fuel_remaining_km': current_position.get('fuel_remaining_km'),
                    'heading': current_position.get('heading'),
                    'timestamp': current_position.get('timestamp', ''),
                }

            result = {
                'object_id': object_id,
                'date': date_str,
                'vehicle': vehicle_info,
                'current_position': cur_pos,
                'odometer_km': current_km,
                'total_distance_km': round(total_distance, 1),
                'total_duration_seconds': total_duration_seconds,
                'total_duration_display': _format_duration(total_duration_seconds),
                'max_speed': round(max_speed_overall),
                'trip_count': len(trip_details),
                'trips': trip_details,
                'fuel_tank_capacity': fuel_tank_capacity,
                'total_fuel_used_liters': total_fuel_used_liters,
                'total_fuel_used_pct': total_fuel_used_pct,
                'avg_fuel_consumption': avg_fuel_consumption,
                'avg_fuel_consumption_pct': avg_fuel_consumption_pct,
                'fuel_per_hour': fuel_per_hour,
                'fuel_per_hour_pct': fuel_per_hour_pct,
                'remaining_drive_seconds': remaining_drive_seconds,
                'remaining_drive_display': _format_duration(remaining_drive_seconds),
                'fuel_chart_data': fuel_chart_data,
                'route_coordinates': route_coordinates,
            }

            return Response(result)

        except FMTrackError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
