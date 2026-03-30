"""
Excel generator for report requests.
Uses openpyxl to produce formatted Excel workbooks.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel(title: str, columns: list, rows: list) -> bytes:
    """
    Generate an Excel workbook and return its contents as bytes.

    Args:
        title: Report title (used as sheet title).
        columns: List of column header strings.
        rows: List of row lists (must match column count).

    Returns:
        bytes: Excel file content.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars

    # ---- Styles ----
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    alt_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    title_font = Font(name='Calibri', bold=True, size=14)
    normal_font = Font(name='Calibri', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Title row ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = center

    # ---- Header row ----
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # ---- Data rows ----
    for row_idx, row in enumerate(rows, start=3):
        fill = alt_fill if (row_idx % 2 == 0) else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.alignment = left
            cell.border = border
            if fill:
                cell.fill = fill

    # ---- Auto-fit column widths ----
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        col_lengths = [len(str(col_name))] + [
            len(str(row[col_idx - 1])) for row in rows if col_idx - 1 < len(row)
        ]
        max_length = max(col_lengths, default=10)
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    # ---- Freeze header ----
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
