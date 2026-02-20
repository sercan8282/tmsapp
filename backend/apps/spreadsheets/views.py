import io
from collections import defaultdict
from datetime import datetime, date
from decimal import Decimal
from django.core.mail import EmailMessage
from django.db.models import Count, Min, Max, Sum, F
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from apps.core.permissions import IsAdminOnly
from apps.core.models import AppSettings
from apps.timetracking.models import TimeEntry
from .models import Spreadsheet
from .serializers import SpreadsheetListSerializer, SpreadsheetDetailSerializer


def safe_str(value):
    """Convert value to safe ASCII string (handle Unicode characters)."""
    if value is None:
        return ''
    s = str(value)
    replacements = {
        '\u0130': 'I', '\u0131': 'i', '\u015e': 'S', '\u015f': 's',
        '\u011e': 'G', '\u011f': 'g', '\u00c7': 'C', '\u00e7': 'c',
        '\u00d6': 'O', '\u00f6': 'o', '\u00dc': 'U', '\u00fc': 'u',
    }
    for char, replacement in replacements.items():
        s = s.replace(char, replacement)
    return s


class SpreadsheetViewSet(viewsets.ModelViewSet):
    """ViewSet for spreadsheet CRUD and actions."""
    permission_classes = [IsAuthenticated, IsAdminOnly]
    filterset_fields = ['bedrijf', 'week_nummer', 'jaar', 'status']
    search_fields = ['naam', 'bedrijf__naam']
    ordering_fields = ['naam', 'week_nummer', 'jaar', 'totaal_factuur', 'updated_at', 'created_at']
    ordering = ['-jaar', '-week_nummer']

    def get_queryset(self):
        return Spreadsheet.objects.select_related('bedrijf', 'created_by').all()

    def get_serializer_class(self):
        if self.action == 'list':
            return SpreadsheetListSerializer
        return SpreadsheetDetailSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Markeer spreadsheet als ingediend."""
        spreadsheet = self.get_object()
        from .models import SpreadsheetStatus
        spreadsheet.status = SpreadsheetStatus.INGEDIEND
        spreadsheet.save()
        serializer = SpreadsheetDetailSerializer(spreadsheet)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def reopen(self, request, pk=None):
        """Zet spreadsheet terug naar concept."""
        spreadsheet = self.get_object()
        from .models import SpreadsheetStatus
        spreadsheet.status = SpreadsheetStatus.CONCEPT
        spreadsheet.save()
        serializer = SpreadsheetDetailSerializer(spreadsheet)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Dupliceer een spreadsheet."""
        original = self.get_object()
        new_sheet = Spreadsheet.objects.create(
            naam=f"{original.naam} (kopie)",
            bedrijf=original.bedrijf,
            week_nummer=original.week_nummer,
            jaar=original.jaar,
            tarief_per_uur=original.tarief_per_uur,
            tarief_per_km=original.tarief_per_km,
            tarief_dot=original.tarief_dot,
            rijen=original.rijen,
            notities=original.notities,
            created_by=request.user,
        )
        serializer = SpreadsheetDetailSerializer(new_sheet)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='available-weeks')
    def available_weeks(self, request):
        """Return last 10 weeks with time entries, grouped per chauffeur.

        Query parameters:
          - jaar (optional): filter by year
        """
        from apps.timetracking.models import TimeEntry

        qs = TimeEntry.objects.select_related('user').all()

        jaar = request.query_params.get('jaar')
        if jaar:
            qs = qs.filter(datum__year=int(jaar))

        # Get distinct week/year combinations with aggregate data
        week_data = (
            qs.values('weeknummer', 'datum__year')
            .annotate(
                count=Count('id'),
                min_datum=Min('datum'),
                max_datum=Max('datum'),
            )
            .order_by('-datum__year', '-weeknummer')[:10]
        )

        result = []
        for wd in week_data:
            wk = wd['weeknummer']
            yr = wd['datum__year']

            # Get per-chauffeur breakdown
            chauffeurs_qs = (
                qs.filter(weeknummer=wk, datum__year=yr)
                .values('user__id', 'user__voornaam', 'user__achternaam')
                .annotate(
                    entries=Count('id'),
                    totaal_km=Sum(F('km_eind') - F('km_start')),
                )
                .order_by('user__voornaam', 'user__achternaam')
            )

            chauffeurs = []
            for ch in chauffeurs_qs:
                naam = f"{ch['user__voornaam']} {ch['user__achternaam']}".strip()
                # Calculate total hours for this chauffeur this week
                ch_entries = qs.filter(
                    weeknummer=wk, datum__year=yr, user_id=ch['user__id']
                )
                total_hours = 0
                for entry in ch_entries:
                    begin = entry.aanvang.hour + entry.aanvang.minute / 60
                    eind = entry.eind.hour + entry.eind.minute / 60
                    if eind < begin:
                        eind += 24
                    pauze = entry.pauze.total_seconds() / 3600
                    total_hours += (eind - begin) - pauze

                chauffeurs.append({
                    'id': str(ch['user__id']),
                    'naam': naam,
                    'entries': ch['entries'],
                    'totaal_uren': round(total_hours, 2),
                    'totaal_km': ch['totaal_km'] or 0,
                })

            result.append({
                'week_nummer': wk,
                'jaar': yr,
                'count': wd['count'],
                'datum_van': wd['min_datum'].strftime('%d-%m-%Y') if wd['min_datum'] else '',
                'datum_tot': wd['max_datum'].strftime('%d-%m-%Y') if wd['max_datum'] else '',
                'chauffeurs': chauffeurs,
            })

        return Response(result)

    @action(detail=False, methods=['get'], url_path='import-time-entries')
    def import_time_entries(self, request):
        """Haal urenregistraties op om te importeren als spreadsheet-rijen.

        Query parameters:
          - week_nummer (required): weeknummer
          - jaar (required): jaar
          - user (optional): UUID van een specifieke chauffeur
        """
        week_nummer = request.query_params.get('week_nummer')
        jaar = request.query_params.get('jaar')
        user_id = request.query_params.get('user')

        if not week_nummer or not jaar:
            return Response(
                {'error': 'week_nummer en jaar zijn verplicht.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        entries = TimeEntry.objects.select_related('user').filter(
            weeknummer=int(week_nummer),
            datum__year=int(jaar),
        ).order_by('datum', 'aanvang')

        if user_id:
            entries = entries.filter(user_id=user_id)

        rijen = []
        for entry in entries:
            # Convert TimeField to decimal hours (e.g. 8:30 -> 8.5)
            begin_decimal = entry.aanvang.hour + entry.aanvang.minute / 60
            eind_decimal = entry.eind.hour + entry.eind.minute / 60
            # Handle overnight: if end < start, add 24
            if eind_decimal < begin_decimal:
                eind_decimal += 24

            # Convert pauze duration to decimal hours
            pauze_hours = entry.pauze.total_seconds() / 3600

            chauffeur_naam = f"{entry.user.voornaam} {entry.user.achternaam}".strip()

            rijen.append({
                'ritnr': entry.ritnummer,
                'volgnummer': '',
                'chauffeur': chauffeur_naam,
                'datum': entry.datum.strftime('%d-%m-%y'),
                'begin_tijd': round(begin_decimal, 2),
                'eind_tijd': round(eind_decimal, 2),
                'pauze': round(pauze_hours, 2),
                'correctie': None,
                'begin_km': entry.km_start,
                'eind_km': entry.km_eind,
                'overnachting': None,
                'overige_kosten': None,
                'time_entry_id': str(entry.id),
            })

        return Response({
            'count': len(rijen),
            'rijen': rijen,
        })

    @action(detail=True, methods=['post'])
    def send_email(self, request, pk=None):
        """E-mail spreadsheet als gestylde XLSX bijlage."""
        spreadsheet = self.get_object()
        emails = request.data.get('emails', [])
        email = request.data.get('email', '')

        if email and not emails:
            emails = [email]
        if not emails:
            return Response(
                {'error': 'Geen e-mailadres opgegeven.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            settings = AppSettings.objects.first()
            if not settings or not settings.smtp_host:
                return Response(
                    {'error': 'SMTP instellingen niet geconfigureerd.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Build styled XLSX
            xlsx_bytes = self._build_xlsx(spreadsheet)

            from django.core.mail.backends.smtp import EmailBackend

            connection = EmailBackend(
                host=safe_str(settings.smtp_host),
                port=settings.smtp_port,
                username=safe_str(settings.smtp_username),
                password=settings.smtp_password,
                use_tls=settings.smtp_use_tls,
                fail_silently=False,
            )

            subject = f"Ritregistratie {spreadsheet.naam} - Week {spreadsheet.week_nummer}/{spreadsheet.jaar}"
            body = (
                f"Bijgevoegd vindt u de ritregistratie:\n\n"
                f"Bedrijf: {spreadsheet.bedrijf.naam}\n"
                f"Week: {spreadsheet.week_nummer} / {spreadsheet.jaar}\n"
                f"Totaal factuur: € {spreadsheet.totaal_factuur:,.2f}\n\n"
                f"Met vriendelijke groet"
            )

            msg = EmailMessage(
                subject=subject,
                body=body,
                from_email=safe_str(settings.smtp_from_email),
                to=emails,
                connection=connection,
            )
            filename = f"ritregistratie_week{spreadsheet.week_nummer}_{spreadsheet.jaar}.xlsx"
            msg.attach(
                filename,
                xlsx_bytes,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            msg.send()

            return Response({
                'message': f'E-mail verzonden naar {", ".join(emails)}',
            })

        except Exception as e:
            return Response(
                {'error': f'E-mail verzenden mislukt: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @staticmethod
    def _fmt_decimal(v):
        """Format decimal value: return empty string for 0/None, number otherwise."""
        if v is None:
            return ''
        v = float(v)
        if v == 0:
            return ''
        return round(v, 2)

    def _build_xlsx(self, spreadsheet):
        """Build a styled XLSX workbook – no background colors, red text for key columns."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = 'Ritregistratie'

        # Style constants
        red_bold = Font(bold=True, color='FF0000', size=10)
        red_font = Font(color='FF0000')
        bold_font = Font(bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right')
        number_fmt = '#,##0.00'
        # Red-text columns: TOTAAL UREN=11, TOTAAL KM=14, TARIEF UUR=15, TARIEF KM=16, DOT=18
        red_cols = {11, 14, 15, 16, 18}

        # Column widths
        letters = 'A B C D E F G H I J K L M N O P Q R S T U'.split()
        col_widths = [8, 12, 12, 14, 12, 9, 9, 9, 9, 12, 14, 10, 10, 12, 15, 15, 12, 10, 16, 17, 14]
        for i, letter in enumerate(letters):
            ws.column_dimensions[letter].width = col_widths[i]

        # Row 1-2: Company header
        ws.merge_cells('A1:D2')
        ws['A1'] = spreadsheet.bedrijf.naam
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(vertical='center')
        ws.merge_cells('E1:G1')
        ws['E1'] = spreadsheet.naam
        ws['E1'].font = Font(bold=True, size=14)
        ws['E1'].alignment = Alignment(vertical='center')
        ws['H1'] = 'Week:'
        ws['H1'].font = Font(size=12)
        ws['H1'].alignment = Alignment(horizontal='right', vertical='center')
        ws['I1'] = spreadsheet.week_nummer
        ws['I1'].font = Font(bold=True, size=14)
        ws.row_dimensions[1].height = 30

        # Row 4: Tariff sub-headers (red text, no bg)
        row4 = 4
        cell = ws.cell(row=row4, column=15, value='TARIEF PER UUR')
        cell.font = red_bold; cell.alignment = center; cell.border = thin_border
        cell = ws.cell(row=row4, column=16, value='TARIEF PER KM')
        cell.font = red_bold; cell.alignment = center; cell.border = thin_border
        cell = ws.cell(row=row4, column=18, value='TARIEF DOT')
        cell.font = red_bold; cell.alignment = center; cell.border = thin_border
        cell = ws.cell(row=row4, column=21, value='totaal factuur')
        cell.font = bold_font; cell.alignment = center; cell.border = thin_border

        # Row 5: Main column headers (bold, red for key cols, no bg)
        row5 = 5
        headers = [
            'WEEK', 'RITNR', 'volgnummer', 'CHAUFFEUR', 'DATUM',
            'BEGIN', 'EIND', 'TOTAAL', 'PAUZE', 'CORRECTIE', 'TOTAAL UREN',
            'BEGIN KM', 'EIND KM', 'TOTAAL KM',
            float(spreadsheet.tarief_per_uur),
            float(spreadsheet.tarief_per_km),
            'totaal',
            float(spreadsheet.tarief_dot),
            'OVERNACHTING', 'OVERIGE KOSTEN', '',
        ]
        for col, val in enumerate(headers, 1):
            cell = ws.cell(row=row5, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center
            if col in red_cols:
                cell.font = red_bold
                if col in (15, 16, 18):
                    cell.number_format = '0.00'
            else:
                cell.font = bold_font

        # Data starts at row 7
        data_start = 7

        for row_idx, rij in enumerate(spreadsheet.rijen):
            begin_tijd = float(rij.get('begin_tijd', 0) or 0)
            eind_tijd = float(rij.get('eind_tijd', 0) or 0)
            pauze = float(rij.get('pauze', 0) or 0)
            correctie = float(rij.get('correctie', 0) or 0)
            begin_km = float(rij.get('begin_km', 0) or 0)
            eind_km = float(rij.get('eind_km', 0) or 0)
            overnachting = float(rij.get('overnachting', 0) or 0)
            overige_kosten = float(rij.get('overige_kosten', 0) or 0)

            r = data_start + row_idx

            # Static input cells
            ws.cell(row=r, column=1, value=spreadsheet.week_nummer)
            ws.cell(row=r, column=2, value=rij.get('ritnr', ''))
            ws.cell(row=r, column=3, value=rij.get('volgnummer', ''))
            ws.cell(row=r, column=4, value=rij.get('chauffeur', ''))
            # Convert datum to a real date so WEEKDAY() works in Excel
            datum_str = rij.get('datum', '')
            datum_val = datum_str
            if datum_str:
                try:
                    datum_val = datetime.strptime(datum_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    try:
                        datum_val = datetime.strptime(datum_str, '%d-%m-%Y').date()
                    except (ValueError, TypeError):
                        pass
            ws.cell(row=r, column=5, value=datum_val or None)
            ws.cell(row=r, column=6, value=begin_tijd or None)       # F - BEGIN
            ws.cell(row=r, column=7, value=eind_tijd or None)        # G - EIND
            ws.cell(row=r, column=9, value=pauze or None)            # I - PAUZE
            ws.cell(row=r, column=10, value=correctie or None)       # J - CORRECTIE
            ws.cell(row=r, column=12, value=int(begin_km) if begin_km else None)  # L
            ws.cell(row=r, column=13, value=int(eind_km) if eind_km else None)    # M
            ws.cell(row=r, column=19, value=overnachting or None)    # S
            ws.cell(row=r, column=20, value=overige_kosten or None)  # T

            # Date format for column E
            if isinstance(datum_val, date):
                ws.cell(row=r, column=5).number_format = 'DD-MM-YYYY'

            # Excel formulas for calculated columns
            ws.cell(row=r, column=8, value=f'=G{r}-F{r}')                                    # H = TOTAAL
            ws.cell(row=r, column=11, value=f'=H{r}-I{r}-J{r}')                               # K = TOTAAL UREN
            ws.cell(row=r, column=14, value=f'=M{r}-L{r}')                                    # N = TOTAAL KM
            ws.cell(row=r, column=15, value=f'=(IF(WEEKDAY(E{r})=7,1.3,1)*K{r})*$O$5')        # O = weekend-toeslag * uren * tarief
            ws.cell(row=r, column=16, value=f'=N{r}*$P$5')                                    # P = tarief km
            ws.cell(row=r, column=17, value=f'=SUM(O{r}:P{r})')                               # Q = subtotaal
            ws.cell(row=r, column=18, value=f'=N{r}*$R$5')                                    # R = DOT
            ws.cell(row=r, column=21, value=f'=SUM(Q{r}:T{r})')                               # U = rij totaal

            # Styling for all columns
            for col in range(1, 22):
                cell = ws.cell(row=r, column=col)
                cell.border = thin_border
                if col in red_cols:
                    cell.font = red_font
                if col in (15, 16, 17, 18, 19, 20, 21):
                    cell.number_format = number_fmt
                    cell.alignment = right_align

        # Empty row + totals with SUM formulas
        last_data_row = data_start + len(spreadsheet.rijen) - 1
        totals_row = data_start + len(spreadsheet.rijen) + 1
        ws.cell(row=totals_row, column=1, value='totaal').font = Font(bold=True, italic=True)
        ws.cell(row=totals_row, column=2, value='factuur').font = Font(bold=True, italic=True)
        ws.cell(row=totals_row, column=3, value='€').font = Font(bold=True, color='FF0000')
        cell = ws.cell(row=totals_row, column=4, value=f'=U{totals_row}')
        cell.font = Font(bold=True, color='FF0000')
        cell.number_format = number_fmt

        col_letters = {15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U'}
        for col in (15, 16, 17, 18, 19, 20, 21):
            letter = col_letters[col]
            cell = ws.cell(
                row=totals_row, column=col,
                value=f'=SUM({letter}{data_start}:{letter}{last_data_row})',
            )
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.number_format = number_fmt

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
