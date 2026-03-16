"""
Service voor het importeren van uren uit Excel bestanden (planbureau).

Excel structuur (Uren E&UTRANS.xlsx):
- Rij 2: headers
- Rij 3+: data
- Kolommen: week, periode, Datum, Ritlijst, Kenteken, km, uurtarief, DOT,
  Geplande vertrektijd, Ingelogd BC, Begintijd Rit, Eindtijd Rit,
  Uren, Pauze, netto uren, uren Factuur, factuur, ...

Kenteken mapping:
  Excel kenteken (bijv. E&UTRANS1) → Vehicle.ritnummer (bijv. EU trans 1)
  → Driver.voertuig → Driver.gekoppelde_gebruiker → User
"""
import logging
from datetime import timedelta, datetime, date, time as dt_time
from decimal import Decimal, InvalidOperation

from django.db import transaction

from apps.fleet.models import Vehicle
from apps.drivers.models import Driver
from .models import ImportBatch, ImportedTimeEntry

logger = logging.getLogger(__name__)


def _parse_time(val):
    """Parse time value from Excel cell - can be datetime, time, or string."""
    if val is None:
        return None
    if isinstance(val, dt_time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ('%H:%M:%S', '%H:%M', '%H.%M'):
            try:
                return datetime.strptime(val, fmt).time()
            except ValueError:
                continue
    return None


def _parse_decimal(val, default=Decimal('0')):
    """Parse decimal from Excel cell."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    if isinstance(val, str):
        val = val.strip().replace(',', '.')
        if not val:
            return default
        try:
            return Decimal(val)
        except InvalidOperation:
            return default
    return default


def _parse_duration(val):
    """Parse pauze duration from Excel cell (time like 0:45 → 45 min)."""
    if val is None:
        return timedelta(minutes=0)
    if isinstance(val, timedelta):
        return val
    if isinstance(val, dt_time):
        return timedelta(hours=val.hour, minutes=val.minute)
    if isinstance(val, datetime):
        return timedelta(hours=val.hour, minutes=val.minute)
    if isinstance(val, (int, float)):
        # Assume minutes
        return timedelta(minutes=int(val))
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return timedelta(minutes=0)
        if ':' in val:
            parts = val.split(':')
            try:
                return timedelta(hours=int(parts[0]), minutes=int(parts[1]))
            except (ValueError, IndexError):
                pass
    return timedelta(minutes=0)


def _build_kenteken_mapping():
    """
    Build mapping from Excel kenteken → (Vehicle, User).
    
    Uses Driver.voertuig FK and Driver.gekoppelde_gebruiker FK
    to resolve Excel kentekens to users.
    
    Only includes drivers with an active (is_active=True) linked user,
    so inactive chauffeurs are skipped during import.
    """
    mapping = {}

    drivers = Driver.objects.select_related(
        'voertuig', 'gekoppelde_gebruiker'
    ).filter(
        voertuig__isnull=False,
        gekoppelde_gebruiker__isnull=False,
        gekoppelde_gebruiker__is_active=True,
    )

    for driver in drivers:
        vehicle = driver.voertuig
        user = driver.gekoppelde_gebruiker

        entry = {
            'vehicle': vehicle,
            'user': user,
            'driver': driver,
        }

        # Store by vehicle ritnummer (fully normalized)
        ritnummer_key = _normalize_kenteken(vehicle.ritnummer)
        if ritnummer_key:
            mapping[ritnummer_key] = entry

        # Also store by vehicle kenteken (fully normalized)
        kenteken_key = _normalize_kenteken(vehicle.kenteken)
        if kenteken_key:
            mapping[kenteken_key] = entry

    return mapping


def _normalize_kenteken(excel_kenteken):
    """Normalize an Excel kenteken for lookup."""
    if not excel_kenteken:
        return ''
    return str(excel_kenteken).lower().replace('&', '').replace(' ', '').replace('-', '')


def check_duplicates_excel(file_obj):
    """
    Quick check: how many rows in this Excel already exist in the DB?
    A duplicate is defined as same datum + kenteken_import (normalized).
    Returns (duplicate_count, total_rows).
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()

    # Collect (datum, kenteken) pairs from Excel
    excel_pairs = set()
    total = 0
    for row in rows:
        if not row or len(row) < 5:
            continue
        datum_val = row[2]
        kenteken_raw = str(row[4]) if row[4] else ''
        if not kenteken_raw.strip():
            continue
        if datum_val is None:
            continue
        if isinstance(datum_val, datetime):
            datum = datum_val.date()
        elif isinstance(datum_val, date):
            datum = datum_val
        elif isinstance(datum_val, str):
            try:
                datum = datetime.strptime(datum_val, '%Y-%m-%d').date()
            except ValueError:
                continue
        else:
            continue
        total += 1
        excel_pairs.add((datum, _normalize_kenteken(kenteken_raw)))

    if not excel_pairs:
        return 0, total

    # Check which pairs already exist
    existing = ImportedTimeEntry.objects.values_list('datum', 'kenteken_import')
    existing_pairs = set()
    for d, k in existing:
        existing_pairs.add((d, _normalize_kenteken(k)))

    duplicates = excel_pairs & existing_pairs
    return len(duplicates), total


def import_excel(file_obj, filename, uploaded_by, overwrite=False, skip_duplicates=False):
    """
    Import uren from an Excel file.
    
    Returns the created ImportBatch with stats.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active

    # Build kenteken → user mapping
    kenteken_map = _build_kenteken_mapping()

    rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()

    batch = ImportBatch.objects.create(
        bestandsnaam=filename,
        geimporteerd_door=uploaded_by,
        totaal_rijen=0,
    )

    entries_to_create = []
    total = 0
    matched = 0
    unmatched = 0

    for row in rows:
        if not row or len(row) < 5:
            continue

        week_val = row[0]
        if week_val is None:
            continue
        try:
            week_str = str(week_val)
            if '-' in week_str:
                weeknummer = int(week_str.split('-')[1])
            else:
                weeknummer = int(week_str)
        except (ValueError, TypeError, IndexError):
            continue

        datum_val = row[2]
        if datum_val is None:
            continue
        if isinstance(datum_val, datetime):
            datum = datum_val.date()
        elif isinstance(datum_val, date):
            datum = datum_val
        elif isinstance(datum_val, str):
            try:
                datum = datetime.strptime(datum_val, '%Y-%m-%d').date()
            except ValueError:
                continue
        else:
            continue

        kenteken_raw = str(row[4]) if row[4] else ''
        if not kenteken_raw.strip():
            continue

        total += 1

        # Look up vehicle/user
        norm_key = _normalize_kenteken(kenteken_raw)
        match = kenteken_map.get(norm_key)

        vehicle = match['vehicle'] if match else None
        user = match['user'] if match else None
        if match:
            matched += 1
        else:
            unmatched += 1

        entry = ImportedTimeEntry(
            batch=batch,
            user=user,
            weeknummer=weeknummer,
            periode=str(row[1] or ''),
            datum=datum,
            ritlijst=str(row[3] or ''),
            kenteken_import=kenteken_raw.strip(),
            km=_parse_decimal(row[5] if len(row) > 5 else None),
            uurtarief=_parse_decimal(row[6] if len(row) > 6 else None),
            dot=str(row[7] or '') if len(row) > 7 else '',
            geplande_vertrektijd=_parse_time(row[8] if len(row) > 8 else None),
            ingelogd_bc=_parse_time(row[9] if len(row) > 9 else None),
            begintijd_rit=_parse_time(row[10] if len(row) > 10 else None),
            eindtijd_rit=_parse_time(row[11] if len(row) > 11 else None),
            uren=_parse_decimal(row[12] if len(row) > 12 else None),
            pauze=_parse_duration(row[13] if len(row) > 13 else None),
            netto_uren=_parse_decimal(row[14] if len(row) > 14 else None),
            uren_factuur=_parse_decimal(row[15] if len(row) > 15 else None),
            factuur_bedrag=_parse_decimal(row[16] if len(row) > 16 else None),
            gekoppeld_voertuig=vehicle,
        )
        entries_to_create.append((entry, norm_key))

    # Handle duplicates: build lookup of existing entries by (datum, normalized kenteken)
    skipped = 0
    if overwrite or skip_duplicates:
        # Collect all (datum, kenteken) pairs from the new entries
        new_dates = set(e.datum for e, _ in entries_to_create)
        existing_entries = ImportedTimeEntry.objects.filter(
            datum__in=new_dates
        ).exclude(batch=batch)

        existing_lookup = {}
        for ex in existing_entries:
            key = (ex.datum, _normalize_kenteken(ex.kenteken_import))
            if key not in existing_lookup:
                existing_lookup[key] = []
            existing_lookup[key].append(ex)

        if overwrite:
            # Delete existing duplicates, then create all new entries
            ids_to_delete = []
            for entry, norm_key in entries_to_create:
                key = (entry.datum, norm_key)
                if key in existing_lookup:
                    ids_to_delete.extend([e.id for e in existing_lookup[key]])
            if ids_to_delete:
                ImportedTimeEntry.objects.filter(id__in=ids_to_delete).delete()
                logger.info(f"Overwrite mode: deleted {len(ids_to_delete)} existing duplicate entries")

        elif skip_duplicates:
            # Filter out entries that already exist
            filtered = []
            for entry, norm_key in entries_to_create:
                key = (entry.datum, norm_key)
                if key in existing_lookup:
                    skipped += 1
                else:
                    filtered.append((entry, norm_key))
            entries_to_create = filtered
            logger.info(f"Skip mode: skipped {skipped} duplicate entries")

    # Extract just the entry objects for bulk_create
    final_entries = [entry for entry, _ in entries_to_create]

    # Bulk create in a transaction
    with transaction.atomic():
        ImportedTimeEntry.objects.bulk_create(final_entries, batch_size=500)
        batch.totaal_rijen = total - skipped
        batch.gekoppeld = matched - skipped if skip_duplicates else matched
        batch.niet_gekoppeld = unmatched
        batch.save()

    logger.info(
        f"Import batch {batch.id}: {total} rows, {matched} matched, {unmatched} unmatched, {skipped} skipped"
    )

    return batch
